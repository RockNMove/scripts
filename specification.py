from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# === Проверка файла на блокировку ===
excel_path = r"C:\Users\Andrey\Desktop\combinepy\combined_py.xlsx"
try:
    with open(excel_path, "r+b"):
        pass  # Файл доступен
except PermissionError:
    print("❌ Файл уже открыт (возможно, в Excel). Закрой его перед запуском скрипта.")
    exit()

# === Шаг 1: Загрузка и подготовка данных ===
df = pd.read_excel(excel_path)

# Принудительно очищаем содержимое в столбцах "+ спецификация", если они есть
for col in ["Родительский продукт + спецификация", "Юнит + спецификация"]:
    if col in df.columns:
        df[col] = ""

# Удаляем все столбцы после "Юнит" (оставляем только столбцы до и включая "Юнит")
if "Юнит" in df.columns:
    last_base_col = df.columns.get_loc("Юнит")
    df = df.iloc[:, :last_base_col + 1]

# Пересоздаём столбцы "+ спецификация" на основе значений из "Родительский продукт", "Юнит" и "Спецификация"
df["Родительский продукт + спецификация"] = df.apply(
    lambda row: "" if pd.isna(row["Родительский продукт"]) or str(row["Родительский продукт"]).strip() == ""
    else (str(row["Родительский продукт"]).strip() + str(row["Спецификация"]).strip()).strip(),
    axis=1
)
df["Юнит + спецификация"] = df.apply(
    lambda row: (str(row["Юнит"]).strip() + str(row["Спецификация"]).strip()).strip(),
    axis=1
)

# === Шаг 2: Построение дерева спецификации ===
all_output = []
processed_indices = set()


def level_to_column(level):
    return f"Level #{level}"


def process_rows(parent_value, level, output_rows):
    # Если parent_value пустое, ищем строки, где "Родительский продукт + спецификация" пустой
    rows = df[df["Родительский продукт + спецификация"] ==
              parent_value] if parent_value else df[df["Родительский продукт + спецификация"] == ""]
    for index, current_row in rows.iterrows():
        if index in processed_indices:
            continue
        processed_indices.add(index)
        row_dict = current_row.to_dict()
        # Записываем значение из "Юнит" в соответствующий столбец уровня, или "_" если отсутствует
        row_dict[level_to_column(level)] = str(current_row["Юнит"]).strip() if pd.notna(
            current_row["Юнит"]) and str(current_row["Юнит"]).strip() != "" else "_"
        output_rows.append(row_dict)
        # Рекурсивно обрабатываем дочерние записи по ключу "Юнит + спецификация"
        process_rows(current_row["Юнит + спецификация"], level + 1, output_rows)


# Запуск обработки от корневых строк — тех, где "Родительский продукт + спецификация" пустой
for index, root in df[df["Родительский продукт + спецификация"] == ""].iterrows():
    if index in processed_indices:
        continue
    processed_indices.add(index)
    row_dict = root.to_dict()
    row_dict[level_to_column(0)] = str(root["Юнит"]).strip() if pd.notna(
        root["Юнит"]) and str(root["Юнит"]).strip() != "" else "_"
    all_output.append(row_dict)
    process_rows(root["Юнит + спецификация"], 1, all_output)

# === Шаг 3: Постобработка уровней ===
spec_df = pd.DataFrame(all_output)

# Определяем максимальный уровень по столбцам, содержащим "Level #"
level_cols = [col for col in spec_df.columns if "Level #" in col]
max_level = max([int(col.split("#")[1]) for col in level_cols]) if level_cols else 0

for i in range(max_level + 1):
    col = level_to_column(i)
    if col not in spec_df.columns:
        spec_df[col] = "_"
    else:
        spec_df[col] = spec_df[col].fillna("_")

# === Шаг 3.5: Обработка подвисших (не вошедших) строк ===
# Определяем индексы строк, которые не были обработаны
not_processed = set(df.index) - processed_indices
if not_processed:
    print("🔔 Подвисшие строки (не включены в дерево спецификации):")
    print(df.loc[list(not_processed)][["Юнит", "Родительский продукт"]])
    print("Жди применения форматов и сохранения файла...")
    # Для каждой не обработанной строки создаём новую запись с Level #0 = "No parent"
    not_proc_rows = []
    for idx in not_processed:
        row = df.loc[idx]
        row_dict = row.to_dict()
        row_dict[level_to_column(0)] = "No parent"
        not_proc_rows.append(row_dict)
    # Преобразуем список подвисших строк в DataFrame
    not_proc_df = pd.DataFrame(not_proc_rows)
    # Поддерживаем формат столбцов Level: если каких-либо столбцов не хватает — заполняем их "_"
    for i in range(max_level + 1):
        col = level_to_column(i)
        if col not in not_proc_df.columns:
            not_proc_df[col] = "_"
        else:
            not_proc_df[col] = not_proc_df[col].fillna("_")
    # Располагаем подвисшие строки в начале итогового DataFrame
    spec_df = pd.concat([not_proc_df, spec_df], ignore_index=True)
else:
    print("✅ Все строки были включены в дерево спецификации.")
    print("Жди применения форматов и сохранения файла...")

# === Шаг 4: Сохранение итогового Excel-файла ===
# Перезаписываем исходный файл
output_path = excel_path  # перезаписываем оригинальный файл
spec_df.to_excel(output_path, index=False)

# === Шаг 5: Оформление Excel-файла с использованием openpyxl ===
wb = load_workbook(output_path)
ws = wb.active

# Фиксация верхней строки (заголовки)
ws.freeze_panes = "A2"

# Установка автофильтров для всех данных
max_col_letter = get_column_letter(ws.max_column)
ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"

# Определяем позиции столбцов "Артикул" и "Юнит"
headers = [cell.value for cell in ws[1]]
art_idx = headers.index("Артикул") + 1 if "Артикул" in headers else None
unit_idx = headers.index("Юнит") + 1 if "Юнит" in headers else None

# Применяем жирные границы:
#  - Слева от "Артикул": для ячеек, непосредственно левее столбца "Артикул"
#  - Справа от "Юнит": для ячеек, непосредственно правее столбца "Юнит"
thick = Side(style="thick")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if art_idx and art_idx > 1:
        row[art_idx - 2].border = Border(right=thick)
    if unit_idx and unit_idx < ws.max_column:
        row[unit_idx].border = Border(left=thick)

# Заливаем серым все ячейки слева от "Артикул"
gray_fill = PatternFill(start_color="B2B2B2", end_color="B2B2B2", fill_type="solid")
if art_idx and art_idx > 1:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=art_idx - 1):
        for cell in row:
            cell.fill = gray_fill

# Добавляем текущую дату/время в одну ячейку — в следующей ячейке после последнего столбца заголовка (только в заголовке)
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws.cell(row=1, column=ws.max_column + 1, value=timestamp)

# Сохраняем итоговый файл
wb.save(output_path)
print(f"\n📁 Итоговый файл сохранён: {output_path}")
