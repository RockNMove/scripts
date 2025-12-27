import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

# Папка с файлами (без учета пути к рабочему столу – замените на полный путь при необходимости)
folder = r"C:\Users\Andrey\Desktop\combinepy"
combined_path = os.path.join(folder, "combined1.xlsx")

# Создаем итоговую книгу
combined_wb = Workbook()
combined_ws = combined_wb.active

header_copied = False

# Проходим по всем файлам в папке, кроме итогового "combined.xlsx"
for fname in os.listdir(folder):
    if not fname.endswith('.xlsx'):
        continue
    if fname.lower().startswith("combined1"):
        continue
    path = os.path.join(folder, fname)
    # Открываем книгу с data_only=True – формулы превращаются в значения
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Удаляем столбцы: сначала H-L, потом A-С (обратите внимание на порядок, чтобы не нарушить индексацию)
    ws.delete_cols(8, 5)    # удаляем столбцы H, I, J, K, L
    ws.delete_cols(1, 3)    # удаляем столбцы A, B, C

    # Копируем заголовки из диапазона A1:H1 (делается один раз – из первого обработанного файла)
    if not header_copied:
        header = [ws.cell(row=1, column=i).value for i in range(1, 9)]
        combined_ws.append(header)
        header_copied = True

    # Функция копирования данных по диапазону (без заголовка)
    for start_letter, end_letter in [("A", "H"), ("J", "Q"), ("S", "Z")]:
        start_col = column_index_from_string(start_letter)
        end_col = column_index_from_string(end_letter)
        # Итерируем по строкам, начиная со второй (заголовок – первая строка)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=start_col, max_col=end_col):
            # Копируем строку, если хотя бы одна ячейка не пуста
            if any(cell.value is not None for cell in row):
                combined_ws.append([cell.value for cell in row])

            # Для каждой строки, скопированной из файла, сразу после добавления в combined_ws
            # запишем имя файла в ячейку столбца I (9-й столбец)
                combined_ws.cell(row=combined_ws.max_row, column=9, value=fname)

    wb.close()

# Сохраняем итоговый файл
combined_wb.save(combined_path)
print("success!")
